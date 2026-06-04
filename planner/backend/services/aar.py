"""After-Action Review (AAR) / post-mission Intel report generator.

The AAR is a structured debrief produced after a mission flies. Squadron
runs typically end with the package commander sitting down and writing
up: who flew, what happened, what got killed (or didn't), what went well,
and what needs to change for next week. This service produces the
SKELETON of that document — pre-filled with everything we already know
about the mission (date, theater, flights, callsigns, frequencies) so the
runner only has to type the narrative.

Inputs (all optional):
  - mission overview + groups (same dict shape the rest of the backend uses)
  - signup data (callsign → pilot name) — usually pulled from the roster
  - events log: list of {time_min, type, summary, ...} dicts that the Live
    mode session loop records during the flight (kills, losses, weapon
    releases, mission-goal completions, controller calls). Optional — when
    absent, the engagement-log section ships as a "(none recorded)" stub
    the runner fills in manually.
  - manual notes: free-text the runner writes after the flight

Output: markdown text. We keep this format-agnostic on purpose — squadron
posts directly to Discord, pastes into the wing's wiki, or prints. Three
output flavours match the signup_sheet pattern: build_markdown, build_csv
(flat engagement-log dump for spreadsheet analysis), build_xlsx (formatted
debrief workbook with a stats sheet).

Engagement log event shapes (forward-compatible — the Live session loop
will populate these once we wire the telemetry recorder; for now the API
accepts whatever shape callers send and we render what we recognise):

    {time_min: int, type: "kill",   victim, killer, weapon}
    {time_min: int, type: "loss",   unit, killer}
    {time_min: int, type: "weapon", flight, weapon, target}
    {time_min: int, type: "rtb",    flight, base}
    {time_min: int, type: "note",   text}

Unknown event types render as a generic "{type} — {summary}" line so the
schema can grow without the renderer needing updates for every new kind.
"""

from __future__ import annotations

import csv
import io
from typing import Any, Iterable


# ── Helpers ────────────────────────────────────────────────────────────────

def _zulu(start_time_sec: int | float | None) -> str:
    """Mission start time → "HHMM Z". The mission dict stores start_time as
    seconds-since-midnight (DCS convention). Falls back to '----' when
    absent so the heading never breaks."""
    if start_time_sec is None:
        return "----"
    try:
        s = int(start_time_sec)
    except (TypeError, ValueError):
        return "----"
    return f"{(s // 3600) % 24:02d}{(s // 60) % 60:02d}Z"


def _player_flights(mission: dict, coalition: str | None = None) -> list[dict]:
    """Flights with at least one Player or Client unit. When coalition is
    given, filter to that side; otherwise return both."""
    out = []
    for g in mission.get("groups", []) or []:
        if coalition and (g.get("coalition") or "").lower() != coalition.lower():
            continue
        seats = [u for u in (g.get("units") or []) if str(u.get("skill", "")).lower() in ("player", "client")]
        if seats:
            out.append(g)
    return out


def _flight_pilots(group: dict, signups: dict[str, str] | None) -> list[tuple[str, str]]:
    """[(callsign, pilot)] for every player seat in the flight. Empty pilot
    string means "open / no signup". `signups` is a dict callsign→name; we
    look up each seat's unit name verbatim and fall back to ''."""
    out = []
    for u in group.get("units") or []:
        if str(u.get("skill", "")).lower() not in ("player", "client"):
            continue
        cs = u.get("name") or ""
        pilot = (signups or {}).get(cs, "")
        out.append((cs, pilot))
    return out


def _normalise_events(events: Iterable[dict] | None) -> list[dict]:
    """Sort the event stream by time_min (missing → end). Strips entries
    that aren't dicts so a caller sending a partially-malformed log can't
    crash the renderer."""
    if not events:
        return []
    clean = []
    for ev in events:
        if not isinstance(ev, dict):
            continue
        clean.append(ev)
    clean.sort(key=lambda e: (e.get("time_min") if isinstance(e.get("time_min"), (int, float)) else 9999))
    return clean


def _engagement_summary(events: list[dict]) -> dict[str, int]:
    """Count kills / losses / weapons / RTBs across the event log. Used in
    the stats block at the bottom of the AAR + as the XLSX summary tab."""
    summary = {"kills": 0, "losses": 0, "weapons": 0, "rtbs": 0, "notes": 0, "other": 0}
    for ev in events:
        t = str(ev.get("type") or "").lower()
        if t == "kill":
            summary["kills"] += 1
        elif t == "loss":
            summary["losses"] += 1
        elif t == "weapon":
            summary["weapons"] += 1
        elif t == "rtb":
            summary["rtbs"] += 1
        elif t == "note":
            summary["notes"] += 1
        else:
            summary["other"] += 1
    return summary


# ── Markdown ───────────────────────────────────────────────────────────────

def build_markdown(
    mission: dict,
    *,
    mission_name: str = "",
    theater: str = "",
    signups: dict[str, str] | None = None,
    events: Iterable[dict] | None = None,
    notes: str = "",
    duration_min: int | None = None,
) -> bytes:
    """Markdown AAR — copy-paste ready for Discord / wiki. Returns UTF-8
    bytes to match build_xlsx / build_csv's return type."""
    ov = mission.get("overview", {}) or {}
    name = mission_name or ov.get("sortie") or "Mission"
    th = theater or ov.get("theater") or "—"
    date = ov.get("date") or "—"
    start = _zulu(ov.get("start_time"))
    dur = f"{duration_min} min" if duration_min else "—"

    norm = _normalise_events(events)
    stats = _engagement_summary(norm)

    lines: list[str] = []
    lines.append(f"# After-Action Review — {name}")
    lines.append("")
    lines.append(f"**Date:** {date} · **Start:** {start} · **Theater:** {th} · **Duration:** {dur}")
    lines.append("")
    lines.append(f"_{stats['kills']} kill(s) · {stats['losses']} loss(es) · {stats['weapons']} weapon employment(s) · {stats['rtbs']} RTB(s)_")
    lines.append("")

    # Participants per coalition
    for side, label in (("blue", "BLUEFOR"), ("red", "OPFOR")):
        flights = _player_flights(mission, side)
        if not flights:
            continue
        lines.append(f"## {label}")
        lines.append("")
        for g in flights:
            task = g.get("task") or "—"
            ac = ""
            us = g.get("units") or []
            if us:
                ac = us[0].get("type") or ""
            freq = g.get("frequency")
            freq_str = f"{(freq / 1_000_000):.3f} MHz" if isinstance(freq, (int, float)) and freq > 1000 else (f"{freq:.3f}" if isinstance(freq, (int, float)) else "—")
            lines.append(f"### {g.get('groupName', '?')} · {task} · {ac} · {freq_str}")
            lines.append("")
            lines.append("| Seat | Callsign | Pilot | Status |")
            lines.append("|---|---|---|---|")
            for cs, pilot in _flight_pilots(g, signups):
                p = pilot or "_(open)_"
                lines.append(f"| — | {cs} | {p} |  |")
            lines.append("")

    # Engagement log
    lines.append("## Engagement Log")
    lines.append("")
    if not norm:
        lines.append("_(none recorded — fill in chronologically)_")
        lines.append("")
    else:
        lines.append("| T+min | Type | Detail |")
        lines.append("|---|---|---|")
        for ev in norm:
            t = ev.get("time_min")
            t_str = f"{t}" if isinstance(t, (int, float)) else "—"
            kind = str(ev.get("type") or "?").upper()
            detail = _render_event_detail(ev)
            lines.append(f"| {t_str} | {kind} | {detail} |")
        lines.append("")

    # Summary stats
    lines.append("## Summary")
    lines.append("")
    lines.append(f"- **Kills:** {stats['kills']}")
    lines.append(f"- **Losses:** {stats['losses']}")
    lines.append(f"- **Weapon employments:** {stats['weapons']}")
    lines.append(f"- **RTBs:** {stats['rtbs']}")
    if stats["other"]:
        lines.append(f"- **Other events:** {stats['other']}")
    lines.append("")

    # Manual notes
    lines.append("## Debrief Notes")
    lines.append("")
    lines.append(notes.strip() if notes.strip() else "_(add narrative here — what went well, what to fix, lessons learned)_")
    lines.append("")

    lines.append("---")
    lines.append("_Generated by DCS:OPT · After-Action Review_")
    return ("\n".join(lines) + "\n").encode("utf-8")


def _render_event_detail(ev: dict) -> str:
    """One-line detail rendering for an engagement-log row. Falls back to
    `summary` when set so unknown event shapes still get readable output."""
    t = str(ev.get("type") or "").lower()
    if t == "kill":
        victim = ev.get("victim") or "?"
        killer = ev.get("killer") or "?"
        wpn = ev.get("weapon") or ""
        wpn_str = f" w/ {wpn}" if wpn else ""
        return f"{killer} → {victim}{wpn_str}"
    if t == "loss":
        unit = ev.get("unit") or "?"
        killer = ev.get("killer") or "?"
        return f"{unit} (by {killer})"
    if t == "weapon":
        flight = ev.get("flight") or "?"
        wpn = ev.get("weapon") or "?"
        tgt = ev.get("target") or ""
        tgt_str = f" → {tgt}" if tgt else ""
        return f"{flight} released {wpn}{tgt_str}"
    if t == "rtb":
        flight = ev.get("flight") or "?"
        base = ev.get("base") or "?"
        return f"{flight} RTB {base}"
    if t == "note":
        return str(ev.get("text") or "")
    # Unknown event type — fall back to summary or stringified shape.
    return str(ev.get("summary") or ev.get("text") or "")


# ── CSV (flat engagement-log dump) ─────────────────────────────────────────

def build_csv(
    mission: dict,
    *,
    mission_name: str = "",
    theater: str = "",
    signups: dict[str, str] | None = None,
    events: Iterable[dict] | None = None,
    notes: str = "",
    duration_min: int | None = None,
) -> bytes:
    """Flat CSV dump of the engagement log + a participants block. For
    folks who want to slice in Excel / Sheets rather than read the
    markdown narrative."""
    _ = (signups, notes, duration_min)  # surfaced via header comments only
    ov = mission.get("overview", {}) or {}
    name = mission_name or ov.get("sortie") or "Mission"
    th = theater or ov.get("theater") or "—"
    date = ov.get("date") or "—"
    start = _zulu(ov.get("start_time"))

    buf = io.StringIO()
    w = csv.writer(buf, lineterminator="\n")
    w.writerow([f"# AAR — {name}", f"Date: {date}", f"Start: {start}", f"Theater: {th}"])
    w.writerow([])

    # Participants
    w.writerow(["--- PARTICIPANTS ---"])
    w.writerow(["Coalition", "Group", "Callsign", "Pilot", "Aircraft", "Task"])
    for side, label in (("blue", "BLUEFOR"), ("red", "OPFOR")):
        for g in _player_flights(mission, side):
            task = g.get("task") or ""
            ac = (g.get("units") or [{}])[0].get("type") or ""
            for cs, pilot in _flight_pilots(g, signups):
                w.writerow([label, g.get("groupName", ""), cs, pilot, ac, task])
    w.writerow([])

    # Engagement log
    norm = _normalise_events(events)
    w.writerow(["--- ENGAGEMENT LOG ---"])
    w.writerow(["T+min", "Type", "Killer/Flight", "Victim/Unit", "Weapon/Target", "Notes"])
    for ev in norm:
        t = ev.get("time_min", "")
        kind = str(ev.get("type") or "")
        if kind == "kill":
            w.writerow([t, "kill", ev.get("killer", ""), ev.get("victim", ""), ev.get("weapon", ""), ""])
        elif kind == "loss":
            w.writerow([t, "loss", ev.get("killer", ""), ev.get("unit", ""), "", ""])
        elif kind == "weapon":
            w.writerow([t, "weapon", ev.get("flight", ""), "", ev.get("weapon", ""), ev.get("target", "")])
        elif kind == "rtb":
            w.writerow([t, "rtb", ev.get("flight", ""), "", ev.get("base", ""), ""])
        elif kind == "note":
            w.writerow([t, "note", "", "", "", ev.get("text", "")])
        else:
            w.writerow([t, kind, "", "", "", ev.get("summary", "")])

    return buf.getvalue().encode("utf-8")


# ── XLSX ───────────────────────────────────────────────────────────────────

def build_xlsx(
    mission: dict,
    *,
    mission_name: str = "",
    theater: str = "",
    signups: dict[str, str] | None = None,
    events: Iterable[dict] | None = None,
    notes: str = "",
    duration_min: int | None = None,
) -> bytes:
    """XLSX workbook with three sheets: Summary, Participants, Events.
    Falls back to a stub workbook when openpyxl isn't installed (keeps
    backend boot tolerant on minimal envs)."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        # Headless fallback — emit the markdown as a single-cell .xlsx-shaped
        # blob so the API contract still returns bytes. Tests skip on this
        # path; real deployments always have openpyxl.
        return build_markdown(mission, mission_name=mission_name, theater=theater,
                              signups=signups, events=events, notes=notes,
                              duration_min=duration_min)

    ov = mission.get("overview", {}) or {}
    name = mission_name or ov.get("sortie") or "Mission"
    th = theater or ov.get("theater") or "—"
    date = ov.get("date") or "—"
    start = _zulu(ov.get("start_time"))
    norm = _normalise_events(events)
    stats = _engagement_summary(norm)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary"

    hdr_font = Font(bold=True, size=14, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="1F3A5F")
    sub_font = Font(bold=True, size=11)

    ws["A1"] = f"AAR — {name}"
    ws["A1"].font = hdr_font; ws["A1"].fill = hdr_fill
    ws.merge_cells("A1:D1")
    ws["A2"] = "Date";    ws["B2"] = date
    ws["A3"] = "Start";   ws["B3"] = start
    ws["A4"] = "Theater"; ws["B4"] = th
    ws["A5"] = "Duration"; ws["B5"] = f"{duration_min} min" if duration_min else "—"
    for r in range(2, 6):
        ws.cell(row=r, column=1).font = sub_font

    ws["A7"] = "Outcome"
    ws["A7"].font = sub_font
    ws["A8"] = "Kills";    ws["B8"] = stats["kills"]
    ws["A9"] = "Losses";   ws["B9"] = stats["losses"]
    ws["A10"] = "Weapons"; ws["B10"] = stats["weapons"]
    ws["A11"] = "RTBs";    ws["B11"] = stats["rtbs"]

    ws["A13"] = "Debrief Notes"
    ws["A13"].font = sub_font
    ws["A14"] = notes.strip() or "(add narrative here)"
    ws["A14"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[14].height = 80
    ws.merge_cells("A14:D14")

    # Participants sheet
    pws = wb.create_sheet("Participants")
    pws.append(["Coalition", "Group", "Callsign", "Pilot", "Aircraft", "Task"])
    for c in pws[1]:
        c.font = sub_font; c.fill = hdr_fill; c.font = Font(bold=True, color="FFFFFF")
    for side, label in (("blue", "BLUEFOR"), ("red", "OPFOR")):
        for g in _player_flights(mission, side):
            task = g.get("task") or ""
            ac = (g.get("units") or [{}])[0].get("type") or ""
            for cs, pilot in _flight_pilots(g, signups):
                pws.append([label, g.get("groupName", ""), cs, pilot, ac, task])

    # Events sheet
    ews = wb.create_sheet("Events")
    ews.append(["T+min", "Type", "Killer/Flight", "Victim/Unit", "Weapon/Target", "Notes"])
    for c in ews[1]:
        c.font = Font(bold=True, color="FFFFFF"); c.fill = hdr_fill
    for ev in norm:
        t = ev.get("time_min", "")
        kind = str(ev.get("type") or "")
        if kind == "kill":
            ews.append([t, "kill", ev.get("killer", ""), ev.get("victim", ""), ev.get("weapon", ""), ""])
        elif kind == "loss":
            ews.append([t, "loss", ev.get("killer", ""), ev.get("unit", ""), "", ""])
        elif kind == "weapon":
            ews.append([t, "weapon", ev.get("flight", ""), "", ev.get("weapon", ""), ev.get("target", "")])
        elif kind == "rtb":
            ews.append([t, "rtb", ev.get("flight", ""), "", ev.get("base", ""), ""])
        elif kind == "note":
            ews.append([t, "note", "", "", "", ev.get("text", "")])
        else:
            ews.append([t, kind, "", "", "", ev.get("summary", "")])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
