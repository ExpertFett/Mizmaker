"""Signup-sheet generator — ATO (Air Tasking Order) style.

Rewrite (v1.19.16) to match the format Fett's squadron actually uses:
flight-block layout (4 rows per flight, one per seat), per-seat MODE-3
transponder codes, staff positions block at the top, glossary on the
right, and coalition section headers between groups.

Layout (XLSX — CSV mirrors the column scheme as a flat dump):

    Row 1   MISSION SIGN UP SHEET    <name>                EST MISSION START  <hhmm>  <date>
    Row 2-3 (mission timing detail, glossary key starts column Y)
    Row 4   POSITION         PILOT/RIO/NFO     ON STATION     MISSION BRIEF
    Row 5-10  Staff positions (Package Commander, ATC, GCI, etc.)
    Row 11  METAR / MAP cells
    Row 17  COLUMN HEADERS: TASK / INTL PACKET / FLIGHT / PILOT / RIO / MODEX
            / MODE 3 / MODE 1 / DEPLOC / ARRLOC / MSNACFT / ORDNANCE / MSNLOC
            / PUSH TIME / ON STATION / OFF STATION / TANKER / TRACK / ALTITUDE
            / PFREQ / ACTYPE / NAME / TRACK / PFREQ / SFREQ / ACTYPE / TASKING
    Row 18+ Coalition labels in column A ("BLUE", "RED"), then flight blocks
            (4 rows per flight; first row has flight-level info, subsequent
            rows fill per-seat callsign + transponder).

Cells the runner fills (INTL PACKET, MODEX, transponder codes, ordnance
detail, tasking statement, etc.) ship empty — we leave them as visible
blank cells so the squadron's planner has space to type. Cells we CAN
auto-derive from the mission (flight callsign, aircraft type, frequency,
TACAN, dep/arr airbase) are pre-populated.

Round-trip note: the column header "Callsign" lives at column C (the
per-seat callsign column the example labels "FLIGHT"). The importer
auto-detects "Callsign" → maps signups onto matching slots. "Pilot"
column matches too. The other columns are informational; the importer
ignores anything it doesn't recognise.
"""

from __future__ import annotations

import csv
import io
from typing import Any


# ── Column layout (1-indexed for openpyxl) ─────────────────────────────────
# Names match the example sheet so users can fill in cells consistently.
# Header at row 17 (visible to users) uses the SHORT versions. The importer
# auto-detects "Callsign" + "Pilot" (sensitive to substring match).
COLUMNS = [
    # (header_text, importer_friendly_alias, width, runner_or_auto)
    ("TASK",              "Task",            14, "auto"),    # group.task
    ("INTL PACKET",       "Packet",          12, "runner"),
    ("Callsign",          "Callsign",        16, "auto"),    # unit.name (one per row)
    ("Pilot",             "Pilot",           20, "signup"),  # blank
    ("RIO/NFO",           "RIO",             18, "signup"),  # blank (2-seaters)
    ("MODEX",             "Modex",           8,  "runner"),
    ("MODE 3",            "Mode3",           8,  "runner"),
    ("MODE 1",            "Mode1",           8,  "runner"),
    ("DEPLOC",            "DepLoc",          12, "auto"),    # waypoint[0] airbase
    ("ARRLOC",            "ArrLoc",          12, "auto"),    # waypoint[-1] airbase
    ("MSNACFT",           "Aircraft",        14, "auto"),    # unit.type
    ("ORDNANCE",          "Ordnance",        18, "runner"),
    ("MSNLOC",            "MsnLoc",          12, "runner"),
    ("PUSH TIME",         "PushTime",        12, "runner"),
    ("ON STATION",        "OnStation",       14, "runner"),
    ("OFF STATION",       "OffStation",      14, "runner"),
    ("TANKER",            "Tanker",          10, "runner"),
    ("TANKER TRACK",      "TankerTrack",     12, "runner"),
    ("ALTITUDE",          "Altitude",        10, "runner"),
    ("PFREQ",             "Freq",            10, "auto"),    # group.frequency
    ("ACTYPE",            "AcType",          12, "runner"),
    ("CONTROL NAME",      "ControlName",     14, "runner"),
    ("CONTROL TRACK",     "ControlTrack",    14, "runner"),
    ("CONTROL PFREQ",     "ControlFreq",     14, "runner"),
    ("CONTROL SFREQ",     "ControlSfreq",    14, "runner"),
    ("CONTROL ACTYPE",    "ControlActype",   14, "runner"),
    ("TASKING STATEMENT", "Tasking",         48, "runner"),
]
HEADERS = [c[0] for c in COLUMNS]
WIDTHS = [c[2] for c in COLUMNS]
HEADER_ROW = 17
BODY_START_ROW = 18

STAFF_POSITIONS = [
    ("PACKAGE COMMANDER",  "T+ 0:00-2:30"),
    ("DEPARTURE ATC",      "T+ 0:00-2:00"),
    ("ARRIVAL ATC",        "T+ 1:30-2:30"),
    ("OPFOR COMMANDER",    "T+ 0:00-2:30"),
    ("OLYMPUS MASTER",     "T+ 0:00-2:30"),
    ("BLUEFOR GCI",        "T+ 0:00-2:30"),
    ("OPFOR GCI",          "T+ 0:00-2:30"),
]

GLOSSARY = [
    ("DEPLOC",      "DEPARTURE LOCATION"),
    ("ARRLOC",      "ARRIVAL LOCATION"),
    ("MSNLOC",      "MISSION LOCATION"),
    ("ON STATION",  "EXPECTED TIME AT MSNLOC"),
    ("OFF STATION", "EXPECTED TIME OFF MSNLOC"),
    ("PFREQ",       "PRIMARY FREQUENCY"),
    ("SFREQ",       "SECONDARY FREQUENCY"),
    ("IVO",         "IN THE VICINITY OF"),
    ("IOT",         "IN ORDER TO"),
    ("FOM",         "FREEDOM OF MANEUVER"),
    ("NET",         "NO EARLIER THAN"),
    ("NLT",         "NO LATER THAN"),
    ("MODEX",       "AIRCRAFT TAIL NUMBER (3-DIGIT)"),
    ("MODE 1",      "IFF MISSION CODE (2-DIGIT)"),
    ("MODE 3",      "IFF SQUAWK (4-DIGIT)"),
]


# ── Slot extraction ────────────────────────────────────────────────────────

def _is_player_slot(unit: dict) -> bool:
    skill = (unit.get("skill") or "").strip().lower()
    return skill in ("player", "client")


def _player_groups(mission_data: dict, coalition: str) -> list[dict]:
    groups = mission_data.get("groups") or []
    out = []
    for g in groups:
        if g.get("coalition") != coalition:
            continue
        if not any(_is_player_slot(u) for u in (g.get("units") or [])):
            continue
        out.append(g)
    return out


def _flight_aux(group: dict) -> dict[str, str]:
    """Auto-derived fields per flight (constants across the flight's seats):
    Task, DepLoc, ArrLoc, Aircraft, Freq, ControlPFreq (TACAN if present)."""
    units = group.get("units") or []
    first = units[0] if units else {}
    waypoints = group.get("waypoints") or []
    def _loc(wp: dict) -> str:
        # Prefer the airbase name when DCS attached one. Fall back to the
        # waypoint NAME only when it's not a generic auto-name like "WP0",
        # "WP1", "Point 4" — those add noise without telling the user
        # anything useful about where the flight starts/ends.
        name = (wp.get("airdromeName") or "").strip()
        if name:
            return name
        wname = (wp.get("waypoint_name") or "").strip()
        if not wname:
            return ""
        upper = wname.upper().replace(" ", "")
        if upper.startswith("WP") and upper[2:].isdigit():
            return ""
        if upper.startswith("POINT") and upper[5:].isdigit():
            return ""
        return wname
    deploc = arrloc = ""
    if waypoints:
        deploc = _loc(waypoints[0])
        arrloc = _loc(waypoints[-1])
    freq = group.get("frequency")
    freq_s = f"{freq:.3f}" if isinstance(freq, (int, float)) and freq else ""
    tacan = group.get("tacan") or {}
    tacan_s = ""
    if isinstance(tacan, dict) and tacan.get("channel"):
        tacan_s = f"{tacan.get('channel')}{tacan.get('band') or ''}"
    # Cruise altitude — highest planned waypoint altitude, as FL (≥18k) or ft.
    alts = [w.get("altitude_m") for w in waypoints
            if isinstance(w.get("altitude_m"), (int, float)) and w.get("altitude_m")]
    alt_s = ""
    if alts:
        ft = max(alts) * 3.28084
        alt_s = f"FL{round(ft / 100):03d}" if ft >= 18000 else f"{round(ft / 500) * 500:,} ft"
    return {
        "Task":        (group.get("task") or "").strip().upper(),
        "DepLoc":      deploc,
        "ArrLoc":      arrloc,
        "Aircraft":    (first.get("type") or "").strip(),
        "Freq":        freq_s,
        "ControlFreq": tacan_s,
        "Altitude":    alt_s,
    }


def _mission_support(mission_data: dict) -> dict[str, str]:
    """Mission-wide support assets the whole package shares — the tanker
    and the AWACS/GCI controller — derived from the mission's AI groups by
    task. Used to pre-fill the TANKER + CONTROL NAME columns (same on every
    flight)."""
    groups = mission_data.get("groups") or []

    def _first(task_kw: str) -> dict | None:
        for g in groups:
            if (g.get("task") or "").strip().lower() == task_kw:
                return g
        return None

    tk = _first("refueling")
    aw = _first("awacs")
    return {
        "Tanker":      (tk.get("groupName") if tk else "") or "",
        "ControlName": (aw.get("groupName") if aw else "") or "",
    }


def _metar(weather: dict) -> str:
    """One-line wx summary from the mission weather, for the METAR cell.
    Defensive — emits only the parts present, '' if the dict is empty."""
    if not isinstance(weather, dict):
        return ""
    parts: list[str] = []
    g = (weather.get("wind") or {}).get("atGround") or {}
    spd, dr = g.get("speed"), g.get("dir")
    if isinstance(spd, (int, float)) and isinstance(dr, (int, float)):
        parts.append(f"{int(round(dr)) % 360:03d}/{round(spd * 1.94384):02d}kt")
    t = weather.get("temperature_c")
    if isinstance(t, (int, float)):
        parts.append(f"{round(t)}°C")
    q = weather.get("qnh_inhg")
    if isinstance(q, (int, float)):
        parts.append(f"QNH {q:.2f}")
    vis = weather.get("visibility_m")
    if isinstance(vis, (int, float)) and vis:
        parts.append(f"vis {round(vis / 1000)}km")
    base, dens = weather.get("clouds_base_m"), weather.get("clouds_density")
    if isinstance(base, (int, float)) and isinstance(dens, (int, float)) and dens:
        cover = "FEW" if dens <= 2 else "SCT" if dens <= 4 else "BKN" if dens <= 7 else "OVC"
        parts.append(f"{cover} {round(base * 3.28084 / 100) * 100:,}ft")
    return "   ".join(parts)


def _per_seat_callsigns(group: dict) -> list[str]:
    """Callsigns for the group's player+client slots, in declared order."""
    return [(u.get("name") or "").strip() for u in (group.get("units") or []) if _is_player_slot(u)]


def _mission_meta(mission_data: dict, *, mission_name: str, theater: str) -> dict[str, str]:
    overview = mission_data.get("overview") or {}
    start = overview.get("start_time") or 0
    try:
        h = int(start) // 3600
        m = (int(start) % 3600) // 60
        zulu = f"{h:02d}{m:02d}"
    except Exception:
        zulu = ""
    return {
        "name":    mission_name or "—",
        "theater": theater or "—",
        "date":    str(overview.get("date") or "—"),
        "zulu":    zulu or "—",
    }


# ── XLSX output (ATO-style) ────────────────────────────────────────────────

def build_xlsx(mission_data: dict, *, mission_name: str = "", theater: str = "") -> bytes:
    """Build the ATO-style signup workbook. Raises ImportError when openpyxl
    is missing — caller surfaces that."""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Signup"

    meta = _mission_meta(mission_data, mission_name=mission_name, theater=theater)
    support = _mission_support(mission_data)
    metar = _metar((mission_data.get("overview") or {}).get("weather") or {})

    bold = Font(bold=True)
    bold_orange = Font(bold=True, color="C75D00")
    thin = Side(border_style="thin", color="888888")
    cell_border = Border(top=thin, bottom=thin, left=thin, right=thin)
    centre = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    head_fill = PatternFill("solid", fgColor="1A2A40")
    signup_fill = PatternFill("solid", fgColor="2A2A1A")  # gentle highlight on PILOT
    section_fill = PatternFill("solid", fgColor="333333")

    # ── 1. Title + mission start ────────────────────────────────────────
    ws.cell(row=1, column=3, value="MISSION SIGN UP SHEET").font = Font(bold=True, size=14, color="C75D00")
    ws.merge_cells(start_row=1, start_column=3, end_row=1, end_column=8)
    ws.cell(row=1, column=10, value=meta["name"]).font = bold
    ws.merge_cells(start_row=1, start_column=10, end_row=1, end_column=18)
    ws.cell(row=1, column=19, value="EST MISSION START").font = bold
    ws.cell(row=1, column=20, value=meta["zulu"]).font = bold_orange
    ws.cell(row=1, column=21, value=meta["date"]).font = bold

    ws.cell(row=2, column=19, value="THEATER").font = bold
    ws.cell(row=2, column=20, value=meta["theater"])

    # ── 2. Staff positions block (rows 4–10) ────────────────────────────
    ws.cell(row=4, column=3, value="POSITION").font = bold
    ws.cell(row=4, column=5, value="PILOT / RIO / NFO").font = bold
    ws.cell(row=4, column=6, value="ON STATION").font = bold
    ws.cell(row=4, column=8, value="MISSION BRIEF").font = bold
    for i, (role, window) in enumerate(STAFF_POSITIONS):
        r = 5 + i
        ws.cell(row=r, column=3, value=role).font = bold
        c = ws.cell(row=r, column=5, value="")
        c.fill = signup_fill
        c.border = cell_border
        ws.cell(row=r, column=6, value=window)
    # Brief / map / METAR cells (text only — runner fills the content)
    ws.cell(row=5, column=8, value="(paste mission brief here)").alignment = left
    ws.merge_cells(start_row=5, start_column=8, end_row=10, end_column=14)
    ws.cell(row=5, column=8).border = cell_border

    ws.cell(row=12, column=16, value="METAR").font = bold
    ws.cell(row=13, column=16, value=metar or "(runner fills)").alignment = left
    ws.merge_cells(start_row=13, start_column=16, end_row=15, end_column=18)
    ws.cell(row=13, column=16).border = cell_border

    ws.cell(row=12, column=14, value="MAP").font = bold
    ws.cell(row=13, column=14, value=meta["theater"])
    ws.merge_cells(start_row=13, start_column=14, end_row=15, end_column=15)

    # ── 3. Glossary (rows 1+, columns Y/Z) ──────────────────────────────
    ws.cell(row=1, column=25, value="KEY").font = bold
    ws.cell(row=1, column=26, value="MEANING").font = bold
    for i, (term, meaning) in enumerate(GLOSSARY):
        ws.cell(row=2 + i, column=25, value=term).font = bold
        ws.cell(row=2 + i, column=26, value=meaning)

    # ── 4. Column headers (row 17) ──────────────────────────────────────
    for i, h in enumerate(HEADERS, start=1):
        c = ws.cell(row=HEADER_ROW, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = head_fill
        c.alignment = centre
        c.border = cell_border

    # ── 5. Flight blocks ────────────────────────────────────────────────
    row = BODY_START_ROW

    def write_section_label(text: str, r: int) -> None:
        c = ws.cell(row=r, column=1, value=text)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = section_fill
        c.alignment = centre
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(HEADERS))

    def write_flight_block(group: dict, start_row: int) -> int:
        aux = _flight_aux(group)
        callsigns = _per_seat_callsigns(group)
        if not callsigns:
            return start_row
        # First seat row: flight-level info + first callsign
        for seat_idx, callsign in enumerate(callsigns):
            r = start_row + seat_idx
            for col_idx, (header, _alias, _w, kind) in enumerate(COLUMNS, start=1):
                cell = ws.cell(row=r, column=col_idx)
                cell.border = cell_border
                cell.alignment = left if header == "TASKING STATEMENT" else centre
                if header == "Callsign":
                    cell.value = callsign
                elif header == "Pilot":
                    cell.fill = signup_fill  # signups go here
                elif seat_idx == 0:  # flight-level fields only on the first row
                    if header == "TASK":
                        cell.value = aux["Task"]
                    elif header == "DEPLOC":
                        cell.value = aux["DepLoc"]
                    elif header == "ARRLOC":
                        cell.value = aux["ArrLoc"]
                    elif header == "MSNACFT":
                        cell.value = aux["Aircraft"]
                    elif header == "PFREQ":
                        cell.value = aux["Freq"]
                    elif header == "CONTROL PFREQ":
                        cell.value = aux["ControlFreq"]
                    elif header == "ALTITUDE":
                        cell.value = aux["Altitude"]
                    elif header == "TANKER":
                        cell.value = support["Tanker"]
                    elif header == "CONTROL NAME":
                        cell.value = support["ControlName"]
        return start_row + len(callsigns)

    for coalition_label, code in [("BLUEFOR", "blue"), ("OPFOR", "red")]:
        groups = _player_groups(mission_data, code)
        if not groups:
            continue
        write_section_label(coalition_label, row)
        row += 1
        for g in groups:
            row = write_flight_block(g, row)

    # ── 6. Column widths + freeze ───────────────────────────────────────
    for i, w in enumerate(WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    # Freeze under header so the column titles stay visible while scrolling.
    ws.freeze_panes = ws.cell(row=BODY_START_ROW, column=1).coordinate

    # ── Mission summary sheet (kept for reference / glance view) ────────
    ws_m = wb.create_sheet("Mission")
    ws_m["A1"] = "MISSION DETAILS"
    ws_m["A1"].font = Font(bold=True, size=14, color="C75D00")
    ws_m.merge_cells("A1:B1")
    summary = [
        ("Mission",          meta["name"]),
        ("Theater",          meta["theater"]),
        ("Date",             meta["date"]),
        ("Mission start (Z)", meta["zulu"]),
        ("Player flights — BLUEFOR", str(len(_player_groups(mission_data, "blue")))),
        ("Player flights — OPFOR",   str(len(_player_groups(mission_data, "red")))),
        ("Player slots total",
         str(sum(len(_per_seat_callsigns(g))
                 for g in _player_groups(mission_data, "blue") + _player_groups(mission_data, "red")))),
    ]
    for i, (k, v) in enumerate(summary, start=3):
        ws_m.cell(row=i, column=1, value=k).font = bold
        ws_m.cell(row=i, column=2, value=v)
    ws_m.column_dimensions["A"].width = 28
    ws_m.column_dimensions["B"].width = 40

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── CSV output ──────────────────────────────────────────────────────────────

def build_csv(mission_data: dict, *, mission_name: str = "", theater: str = "") -> bytes:
    """CSV mirror of the ATO layout — same columns, flat row dump. Comment
    block at top carries mission metadata so the file still parses cleanly
    in Excel / Sheets but the runner sees context when they open it raw."""
    meta = _mission_meta(mission_data, mission_name=mission_name, theater=theater)
    buf = io.StringIO()
    buf.write(f"# Mission: {meta['name']}\n")
    buf.write(f"# Theater: {meta['theater']}\n")
    buf.write(f"# Date:    {meta['date']}\n")
    buf.write(f"# Start:   {meta['zulu']}Z\n")
    buf.write("#\n")
    buf.write("# Glossary:\n")
    for term, meaning in GLOSSARY:
        buf.write(f"#   {term:<12} {meaning}\n")
    buf.write("\n")

    support = _mission_support(mission_data)
    w = csv.writer(buf)
    w.writerow(HEADERS)
    for coalition_label, code in [("BLUEFOR", "blue"), ("OPFOR", "red")]:
        groups = _player_groups(mission_data, code)
        if not groups:
            continue
        # Section row (rest of columns blank).
        w.writerow([f"--- {coalition_label} ---"] + [""] * (len(HEADERS) - 1))
        for g in groups:
            aux = _flight_aux(g)
            callsigns = _per_seat_callsigns(g)
            for seat_idx, callsign in enumerate(callsigns):
                row = ["" for _ in HEADERS]
                # Per-row callsign / pilot signup blank
                idx = {h: i for i, h in enumerate(HEADERS)}
                row[idx["Callsign"]] = callsign
                if seat_idx == 0:
                    row[idx["TASK"]] = aux["Task"]
                    row[idx["DEPLOC"]] = aux["DepLoc"]
                    row[idx["ARRLOC"]] = aux["ArrLoc"]
                    row[idx["MSNACFT"]] = aux["Aircraft"]
                    row[idx["PFREQ"]] = aux["Freq"]
                    row[idx["CONTROL PFREQ"]] = aux["ControlFreq"]
                    row[idx["ALTITUDE"]] = aux["Altitude"]
                    row[idx["TANKER"]] = support["Tanker"]
                    row[idx["CONTROL NAME"]] = support["ControlName"]
                w.writerow(row)
    return buf.getvalue().encode("utf-8")


# ── Markdown output (Discord-friendly) ─────────────────────────────────────

def build_markdown(mission_data: dict, *, mission_name: str = "", theater: str = "") -> bytes:
    """Compact Markdown table for Discord posts — drops the operational
    columns (TANKER / CONTROL / TASKING) that don't read well at chat width.
    Pilots see: Task / Flight / Callsign / Pilot / Aircraft / Freq."""
    meta = _mission_meta(mission_data, mission_name=mission_name, theater=theater)
    parts: list[str] = []
    parts.append(f"# {meta['name']}")
    parts.append("")
    parts.append(f"**Theater:** {meta['theater']} · **Date:** {meta['date']} · **Start (Z):** {meta['zulu']}")
    parts.append("")
    parts.append("Sign up by replacing `_(open)_` with your name + callsign.")
    parts.append("")
    for coalition_label, code in [("BLUEFOR", "blue"), ("OPFOR", "red")]:
        groups = _player_groups(mission_data, code)
        if not groups:
            continue
        parts.append(f"## {coalition_label}")
        parts.append("")
        for g in groups:
            aux = _flight_aux(g)
            callsigns = _per_seat_callsigns(g)
            if not callsigns:
                continue
            parts.append(f"**{aux['Task']} · {callsigns[0]} flight · {aux['Aircraft']}** — freq {aux['Freq'] or '—'}")
            parts.append("")
            parts.append("| Seat | Callsign | Pilot |")
            parts.append("|------|----------|-------|")
            for i, cs in enumerate(callsigns, start=1):
                parts.append(f"| {i} | {cs} | _(open)_ |")
            parts.append("")
    return ("\n".join(parts) + "\n").encode("utf-8")
