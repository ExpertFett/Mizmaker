"""Tests for the After-Action Review generator (services/aar.py).

Validates:
  - Markdown skeleton renders with the expected sections regardless of
    whether an events log is provided
  - Engagement-log rendering handles every known event type + unknown
    types (forward-compat)
  - CSV emits PARTICIPANTS + ENGAGEMENT LOG section blocks
  - XLSX produces three sheets (Summary / Participants / Events) with the
    summary stats populated from the event count
  - Empty / minimal inputs don't crash the renderer (defensive)
"""

from __future__ import annotations

import csv
import io

import pytest


def _mission() -> dict:
    return {
        "overview": {"date": "2026-06-04", "start_time": 30600, "sortie": "Op RAINMAKER", "theater": "PersianGulf"},
        "groups": [
            {
                "groupName": "Uzi 1", "coalition": "blue", "task": "CAS",
                "frequency": 305_000_000,
                "waypoints": [{"airdromeName": "Al Dhafra"}],
                "units": [
                    {"name": "Uzi 1-1", "type": "FA-18C_hornet", "skill": "Player"},
                    {"name": "Uzi 1-2", "type": "FA-18C_hornet", "skill": "Client"},
                ],
            },
            {
                "groupName": "Springfield 1", "coalition": "blue", "task": "CAP",
                "frequency": 264_000_000,
                "waypoints": [],
                "units": [
                    {"name": "Springfield 1-1", "type": "F-16C_50", "skill": "Player"},
                ],
            },
            {
                "groupName": "Hammer 1", "coalition": "red", "task": "CAP",
                "frequency": 251_000_000,
                "units": [
                    {"name": "Hammer 1-1", "type": "MiG-29S", "skill": "Client"},
                ],
            },
        ],
    }


def _events() -> list[dict]:
    return [
        {"time_min": 15, "type": "weapon", "flight": "Uzi 1-1", "weapon": "GBU-12", "target": "SA-6"},
        {"time_min": 18, "type": "kill",   "killer": "Springfield 1-1", "victim": "Hammer 1-1", "weapon": "AIM-120C"},
        {"time_min": 22, "type": "loss",   "unit": "Uzi 1-2", "killer": "MANPADS"},
        {"time_min": 30, "type": "rtb",    "flight": "Uzi 1-1", "base": "Al Dhafra"},
        {"time_min": 32, "type": "note",   "text": "Tanker offline at AAR1"},
        # An unknown type — renderer should keep going.
        {"time_min": 35, "type": "weird",  "summary": "wing-tip vapor for the bubbas"},
    ]


# ── Helpers ────────────────────────────────────────────────────────────────

class TestHelpers:
    def test_zulu_formats_seconds(self):
        from services.aar import _zulu
        assert _zulu(30600) == "0830Z"
        assert _zulu(0) == "0000Z"
        assert _zulu(None) == "----"
        assert _zulu("garbage") == "----"

    def test_player_flights_filters_by_coalition(self):
        from services.aar import _player_flights
        m = _mission()
        blue = _player_flights(m, "blue")
        red = _player_flights(m, "red")
        assert [g["groupName"] for g in blue] == ["Uzi 1", "Springfield 1"]
        assert [g["groupName"] for g in red] == ["Hammer 1"]

    def test_normalise_events_sorts_by_time(self):
        from services.aar import _normalise_events
        n = _normalise_events([
            {"time_min": 30, "type": "rtb"},
            {"time_min": 5,  "type": "kill"},
            {"time_min": 20, "type": "weapon"},
        ])
        assert [e["time_min"] for e in n] == [5, 20, 30]

    def test_normalise_events_filters_non_dicts(self):
        from services.aar import _normalise_events
        n = _normalise_events(["nope", 5, None, {"time_min": 1, "type": "note"}])
        assert len(n) == 1

    def test_engagement_summary_counts(self):
        from services.aar import _engagement_summary, _normalise_events
        s = _engagement_summary(_normalise_events(_events()))
        assert s["kills"] == 1
        assert s["losses"] == 1
        assert s["weapons"] == 1
        assert s["rtbs"] == 1
        assert s["notes"] == 1
        assert s["other"] == 1


# ── Markdown ───────────────────────────────────────────────────────────────

class TestMarkdown:
    def test_renders_header_and_sections(self):
        from services.aar import build_markdown
        text = build_markdown(_mission(), mission_name="Op RAINMAKER",
                              theater="PersianGulf", duration_min=75).decode("utf-8")
        assert "# After-Action Review — Op RAINMAKER" in text
        assert "PersianGulf" in text
        assert "75 min" in text
        assert "## BLUEFOR" in text
        assert "## OPFOR" in text
        assert "## Engagement Log" in text
        assert "## Summary" in text
        assert "## Debrief Notes" in text

    def test_no_events_shows_placeholder(self):
        from services.aar import build_markdown
        text = build_markdown(_mission()).decode("utf-8")
        assert "_(none recorded — fill in chronologically)_" in text

    def test_events_render_each_known_type(self):
        from services.aar import build_markdown
        text = build_markdown(_mission(), events=_events()).decode("utf-8")
        # Kill line
        assert "Springfield 1-1 → Hammer 1-1" in text
        assert "AIM-120C" in text
        # Loss line
        assert "Uzi 1-2 (by MANPADS)" in text
        # Weapon line
        assert "Uzi 1-1 released GBU-12 → SA-6" in text
        # RTB
        assert "RTB Al Dhafra" in text
        # Note
        assert "Tanker offline at AAR1" in text

    def test_signups_render_pilot_names(self):
        from services.aar import build_markdown
        signups = {"Uzi 1-1": "Fett", "Springfield 1-1": "Maverick"}
        text = build_markdown(_mission(), signups=signups).decode("utf-8")
        assert "| Fett |" in text
        assert "| Maverick |" in text
        # Unfilled seats show the open placeholder
        assert "_(open)_" in text

    def test_unknown_event_type_does_not_crash(self):
        from services.aar import build_markdown
        text = build_markdown(_mission(), events=[
            {"time_min": 1, "type": "spinach", "summary": "popeye called"},
        ]).decode("utf-8")
        assert "SPINACH" in text  # Type column uppercases the kind
        assert "popeye called" in text

    def test_manual_notes_block(self):
        from services.aar import build_markdown
        text = build_markdown(_mission(), notes="Smooth strike, no losses on egress.").decode("utf-8")
        assert "Smooth strike, no losses on egress." in text

    def test_summary_stats_zero_when_no_events(self):
        from services.aar import build_markdown
        text = build_markdown(_mission()).decode("utf-8")
        assert "**Kills:** 0" in text
        assert "**Losses:** 0" in text


# ── CSV ────────────────────────────────────────────────────────────────────

class TestCsv:
    def test_csv_includes_participants_and_events_sections(self):
        from services.aar import build_csv
        text = build_csv(_mission(), events=_events()).decode("utf-8")
        assert "--- PARTICIPANTS ---" in text
        assert "--- ENGAGEMENT LOG ---" in text

    def test_csv_participants_one_row_per_seat(self):
        from services.aar import build_csv
        text = build_csv(_mission(), signups={"Uzi 1-1": "Fett"}).decode("utf-8")
        lines = [l for l in text.splitlines() if l]
        reader = csv.reader(lines)
        rows = list(reader)
        # Find rows starting with BLUEFOR/OPFOR
        seats = [r for r in rows if r and r[0] in ("BLUEFOR", "OPFOR")]
        # Uzi 1 (2 seats) + Springfield 1 (1 seat) + Hammer 1 (1 seat) = 4
        assert len(seats) == 4
        # Pilot column for Uzi 1-1 = "Fett"
        fett_row = [r for r in seats if r[2] == "Uzi 1-1"][0]
        assert fett_row[3] == "Fett"

    def test_csv_events_rows_match_count(self):
        from services.aar import build_csv
        text = build_csv(_mission(), events=_events()).decode("utf-8")
        lines = [l for l in text.splitlines() if l]
        reader = csv.reader(lines)
        rows = list(reader)
        # Find rows after the event header
        try:
            idx = next(i for i, r in enumerate(rows) if r and r[0] == "T+min")
        except StopIteration:
            pytest.fail("Event header row missing")
        event_rows = rows[idx + 1:]
        # 6 events in our fixture
        assert len(event_rows) == 6


# ── XLSX ───────────────────────────────────────────────────────────────────

class TestXlsx:
    def test_xlsx_three_sheets(self):
        try:
            import openpyxl  # noqa: F401
        except ImportError:
            pytest.skip("openpyxl missing")
        from services.aar import build_xlsx
        data = build_xlsx(_mission(), mission_name="Op X", events=_events())
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(data))
        assert "Summary" in wb.sheetnames
        assert "Participants" in wb.sheetnames
        assert "Events" in wb.sheetnames

    def test_xlsx_summary_stats(self):
        try:
            import openpyxl  # noqa: F401
        except ImportError:
            pytest.skip("openpyxl missing")
        from services.aar import build_xlsx
        data = build_xlsx(_mission(), events=_events())
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(data))
        ws = wb["Summary"]
        # Kills row (A8/B8)
        assert ws["A8"].value == "Kills"
        assert ws["B8"].value == 1
        assert ws["A9"].value == "Losses"
        assert ws["B9"].value == 1

    def test_xlsx_participants_rows(self):
        try:
            import openpyxl  # noqa: F401
        except ImportError:
            pytest.skip("openpyxl missing")
        from services.aar import build_xlsx
        data = build_xlsx(_mission(), signups={"Uzi 1-1": "Fett"})
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(data))
        ws = wb["Participants"]
        rows = list(ws.iter_rows(values_only=True))
        # Row 0 is header; expect 4 seats after that
        assert rows[0][0] == "Coalition"
        body = rows[1:]
        assert len(body) == 4
        uzi = [r for r in body if r[2] == "Uzi 1-1"][0]
        assert uzi[3] == "Fett"

    def test_xlsx_events_rows(self):
        try:
            import openpyxl  # noqa: F401
        except ImportError:
            pytest.skip("openpyxl missing")
        from services.aar import build_xlsx
        data = build_xlsx(_mission(), events=_events())
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(data))
        ws = wb["Events"]
        rows = list(ws.iter_rows(values_only=True))
        # Header + 6 events
        assert len(rows) == 7
        assert rows[0][0] == "T+min"


# ── Edge cases ─────────────────────────────────────────────────────────────

class TestEdgeCases:
    def test_empty_mission_renders(self):
        from services.aar import build_markdown
        text = build_markdown({}, mission_name="Empty").decode("utf-8")
        assert "# After-Action Review — Empty" in text
        # No coalition headers when no flights
        assert "## BLUEFOR" not in text
        assert "## OPFOR" not in text

    def test_no_signups_marks_all_seats_open(self):
        from services.aar import build_markdown
        text = build_markdown(_mission()).decode("utf-8")
        # All 4 player seats should be _(open)_
        assert text.count("_(open)_") == 4

    def test_none_events_treated_as_empty(self):
        from services.aar import build_markdown
        text = build_markdown(_mission(), events=None).decode("utf-8")
        assert "_(none recorded" in text
