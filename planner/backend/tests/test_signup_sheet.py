"""Tests for the ATO-style signup sheet generator (services/signup_sheet.py).

Validates that:
  - Player slots are extracted per coalition with correct seat sequencing
  - All three formats (XLSX / CSV / MD) emit the new flight-block layout
  - "Callsign" header still matches the Roster importer's auto-detect
    (so a filled XLSX can be uploaded back via the existing flow)
  - Coalition section labels (BLUEFOR / OPFOR) appear in the body
  - Auto-derived flight fields (Task, MSNACFT, PFREQ) land on the first
    seat row only
"""

from __future__ import annotations

import csv
import io

import pytest


def _two_flight_mission() -> dict:
    """Synthetic mission — one all-player blue strike flight (4 seats),
    one mixed blue SEAD flight (1 player + 1 AI), one red CAP (1 client)."""
    return {
        "overview": {"date": "2018-06-01", "start_time": 30600, "sortie": ""},  # 8:30Z
        "groups": [
            {
                "groupName": "Camelot 1", "coalition": "blue", "task": "CAS",
                "frequency": 305.0, "tacan": {"channel": 73, "band": "X"},
                "waypoints": [
                    {"airdromeName": "Nellis"}, {}, {}, {"airdromeName": "Nellis"},
                ],
                "units": [
                    {"name": "Camelot 1-1", "type": "FA-18C_hornet", "skill": "Player"},
                    {"name": "Camelot 1-2", "type": "FA-18C_hornet", "skill": "Client"},
                    {"name": "Camelot 1-3", "type": "FA-18C_hornet", "skill": "Player"},
                    {"name": "Camelot 1-4", "type": "FA-18C_hornet", "skill": "Player"},
                ],
            },
            {
                "groupName": "Bengal 1", "coalition": "blue", "task": "SEAD",
                "frequency": 332.1, "tacan": None,
                "waypoints": [
                    {"airdromeName": "Nellis"}, {}, {"airdromeName": "Nellis"},
                ],
                "units": [
                    {"name": "Bengal 1-1", "type": "F-16C_50", "skill": "Player"},
                    {"name": "Bengal 1-2", "type": "F-16C_50", "skill": "High"},  # AI
                ],
            },
            {
                "groupName": "Jackal 1", "coalition": "red", "task": "CAP",
                "frequency": 264.0, "tacan": None,
                "waypoints": [],
                "units": [
                    {"name": "Jackal 1-1", "type": "MiG-29S", "skill": "Client"},
                ],
            },
        ],
    }


# ── Slot extraction ────────────────────────────────────────────────────────

class TestSlotExtraction:
    def test_player_groups_per_coalition(self):
        from services.signup_sheet import _player_groups
        m = _two_flight_mission()
        blue = _player_groups(m, "blue")
        red = _player_groups(m, "red")
        assert [g["groupName"] for g in blue] == ["Camelot 1", "Bengal 1"]
        assert [g["groupName"] for g in red] == ["Jackal 1"]

    def test_per_seat_callsigns(self):
        from services.signup_sheet import _per_seat_callsigns
        m = _two_flight_mission()
        cs = _per_seat_callsigns(m["groups"][0])  # Camelot 1
        # All 4 (all Player/Client)
        assert cs == ["Camelot 1-1", "Camelot 1-2", "Camelot 1-3", "Camelot 1-4"]

    def test_skill_high_excluded(self):
        from services.signup_sheet import _per_seat_callsigns
        m = _two_flight_mission()
        cs = _per_seat_callsigns(m["groups"][1])  # Bengal 1 — 1 player + 1 AI
        assert cs == ["Bengal 1-1"]

    def test_flight_aux_pulls_task_aircraft_freq(self):
        from services.signup_sheet import _flight_aux
        m = _two_flight_mission()
        aux = _flight_aux(m["groups"][0])
        assert aux["Task"] == "CAS"
        assert aux["Aircraft"] == "FA-18C_hornet"
        assert aux["Freq"] == "305.000"
        assert aux["DepLoc"] == "Nellis"
        assert aux["ArrLoc"] == "Nellis"
        assert aux["ControlFreq"] == "73X"


# ── CSV ────────────────────────────────────────────────────────────────────

class TestCsv:
    def test_csv_header_includes_callsign_and_pilot(self):
        """The importer's auto-detect looks for 'callsign' + 'pilot' (case-
        insensitive substring). Both columns must be present so a filled
        CSV round-trips through RosterTab.upload."""
        from services.signup_sheet import build_csv
        data = build_csv(_two_flight_mission(), mission_name="Test", theater="NTTR")
        text = data.decode("utf-8")
        # Find the header line — first line that isn't a comment/blank.
        lines = [l for l in text.splitlines() if l.strip() and not l.startswith("#")]
        headers = next(csv.reader([lines[0]]))
        lower = [h.lower() for h in headers]
        assert any("callsign" in h for h in lower)
        assert any("pilot" in h for h in lower)

    def test_csv_emits_coalition_section_rows(self):
        from services.signup_sheet import build_csv
        text = build_csv(_two_flight_mission(), mission_name="X", theater="Y").decode("utf-8")
        assert "--- BLUEFOR ---" in text
        assert "--- OPFOR ---" in text

    def test_csv_one_row_per_player_seat(self):
        from services.signup_sheet import build_csv
        text = build_csv(_two_flight_mission(), mission_name="X", theater="Y").decode("utf-8")
        lines = [l for l in text.splitlines() if l.strip() and not l.startswith("#")]
        reader = csv.reader(lines)
        headers = next(reader)
        cs_idx = headers.index("Callsign")
        callsigns = [r[cs_idx] for r in reader if cs_idx < len(r) and r[cs_idx] and not r[cs_idx].startswith("---")]
        # 4 (Camelot) + 1 (Bengal player) + 1 (Jackal) = 6
        assert len(callsigns) == 6
        assert "Camelot 1-1" in callsigns
        assert "Jackal 1-1" in callsigns

    def test_csv_task_and_aircraft_only_on_first_seat(self):
        """Task/Aircraft/Freq are flight-level — should land on the first
        seat row only, with the following seats empty in those columns."""
        from services.signup_sheet import build_csv
        text = build_csv(_two_flight_mission(), mission_name="X", theater="Y").decode("utf-8")
        lines = [l for l in text.splitlines() if l.strip() and not l.startswith("#")]
        reader = csv.reader(lines)
        headers = next(reader)
        rows = [r for r in reader if r and r[0] != "" or len(r) > 1]
        # Find Camelot block (after BLUEFOR section)
        body = [r for r in rows if not r[0].startswith("---")]
        # Camelot's 4 rows are the first 4 in body
        cs_idx = headers.index("Callsign")
        task_idx = headers.index("TASK")
        camelot = [r for r in body if cs_idx < len(r) and r[cs_idx].startswith("Camelot")]
        assert camelot[0][task_idx] == "CAS"  # first seat carries it
        # subsequent seats: task column empty
        for r in camelot[1:]:
            assert r[task_idx] == ""


# ── Markdown ───────────────────────────────────────────────────────────────

class TestMarkdown:
    def test_markdown_per_flight_subheaders(self):
        from services.signup_sheet import build_markdown
        text = build_markdown(_two_flight_mission(), mission_name="X", theater="NTTR").decode("utf-8")
        # Section headers
        assert "## BLUEFOR" in text
        assert "## OPFOR" in text
        # Per-flight title block uses Task · lead-callsign · aircraft
        assert "CAS · Camelot 1-1 flight · FA-18C_hornet" in text
        assert "SEAD · Bengal 1-1 flight · F-16C_50" in text

    def test_markdown_open_placeholder(self):
        from services.signup_sheet import build_markdown
        text = build_markdown(_two_flight_mission(), mission_name="X", theater="Y").decode("utf-8")
        # Empty pilot cells render as _(open)_
        assert "_(open)_" in text

    def test_markdown_includes_seat_table(self):
        from services.signup_sheet import build_markdown
        text = build_markdown(_two_flight_mission(), mission_name="X", theater="Y").decode("utf-8")
        # Camelot's 4 seats appear as rows in a table.
        for cs in ("Camelot 1-1", "Camelot 1-2", "Camelot 1-3", "Camelot 1-4"):
            assert f"| {cs} |" in text


# ── XLSX ───────────────────────────────────────────────────────────────────

class TestXlsx:
    def test_xlsx_has_signup_and_mission_sheets(self):
        try:
            import openpyxl  # noqa: F401
        except ImportError:
            pytest.skip("openpyxl missing")
        from services.signup_sheet import build_xlsx
        data = build_xlsx(_two_flight_mission(), mission_name="Test Op", theater="NTTR")
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(data))
        assert "Signup" in wb.sheetnames
        assert "Mission" in wb.sheetnames

    def test_xlsx_column_headers_at_row_17(self):
        try:
            import openpyxl  # noqa: F401
        except ImportError:
            pytest.skip("openpyxl missing")
        from services.signup_sheet import build_xlsx, HEADERS
        data = build_xlsx(_two_flight_mission(), mission_name="X", theater="Y")
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(data))
        ws = wb["Signup"]
        actual = [ws.cell(row=17, column=i + 1).value for i in range(len(HEADERS))]
        assert actual == HEADERS

    def test_xlsx_flight_blocks_4_rows_per_camelot(self):
        try:
            import openpyxl  # noqa: F401
        except ImportError:
            pytest.skip("openpyxl missing")
        from services.signup_sheet import build_xlsx, HEADERS
        data = build_xlsx(_two_flight_mission(), mission_name="X", theater="Y")
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(data))
        ws = wb["Signup"]
        cs_col = HEADERS.index("Callsign") + 1
        # Body starts row 18; row 18 should be BLUEFOR section label,
        # rows 19-22 = Camelot 1-1..1-4.
        bluefor = ws.cell(row=18, column=1).value
        assert bluefor == "BLUEFOR"
        callsigns = [ws.cell(row=19 + i, column=cs_col).value for i in range(4)]
        assert callsigns == ["Camelot 1-1", "Camelot 1-2", "Camelot 1-3", "Camelot 1-4"]

    def test_xlsx_flight_level_fields_first_seat_only(self):
        try:
            import openpyxl  # noqa: F401
        except ImportError:
            pytest.skip("openpyxl missing")
        from services.signup_sheet import build_xlsx, HEADERS
        data = build_xlsx(_two_flight_mission(), mission_name="X", theater="Y")
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(data))
        ws = wb["Signup"]
        task_col = HEADERS.index("TASK") + 1
        ac_col = HEADERS.index("MSNACFT") + 1
        # Camelot first row = 19
        assert ws.cell(row=19, column=task_col).value == "CAS"
        assert ws.cell(row=19, column=ac_col).value == "FA-18C_hornet"
        # Camelot seat 2 (row 20) — flight-level fields empty
        assert ws.cell(row=20, column=task_col).value in (None, "")
        assert ws.cell(row=20, column=ac_col).value in (None, "")
